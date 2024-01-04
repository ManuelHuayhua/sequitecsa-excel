from flask import Flask, render_template, request, send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO
from datetime import datetime


app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'  # Cambia esto por una clave segura

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process_name', methods=['POST'])
def process_name():
    cliente = request.form['cliente']
    proyecto = request.form['proyecto']
    contacto = request.form['contacto'] 
    cargo = request.form['cargo']
    
    return render_template('cliente.html', cliente=cliente, proyecto=proyecto, contacto=contacto, cargo=cargo)

@app.route('/download_excel/<cliente>/<proyecto>/<contacto>/<cargo>')
def download_excel(cliente, proyecto, contacto,cargo):
    # Crear un libro de Excel y una hoja de cálculo
    wb = Workbook()
    ws = wb.active

    # Establecer estilos y formato en la hoja de cálculo
    for col in range(1, 22):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 10

    for row in range(1, 5):
        ws.row_dimensions[row].height = 20

    for col_letter in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 7

    for col_letter in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col_letter].width = 8

    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=17):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))

    ws['A1'] = "Hola Mundo"
    ws.merge_cells('A1:D4')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['E1'] = "Gestión Comercial"
    ws.merge_cells('E1:O1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['E2'] = "FORMATO"
    ws.merge_cells('E2:O2')
    ws['E2'].alignment = Alignment(horizontal='center', vertical='center')

    ws['E3'] = "ENCUESTA DE SATISFACCIÓN DEL CLIENTE"
    ws.merge_cells('E3:O4')
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=4, min_col=5, max_col=15):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws['P1:Q4'][0]:
        cell.border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))

    ws['P1'] = "CÓDIGO"
    ws['Q1'] = "SQ-COM-F-003"
    ws.column_dimensions['Q'].width = 15
    
    ws['P2'] = "F. DE /n APROBACIÓN"
    ws['Q2'] = "1/7/2023"
    ws['P3'] = "VERSION"
    ws['Q3'] = "01"
    ws['P4'] = "PÁGINA"
    ws['Q4'] = "1/1"

    for cell in ws['P1:Q4'][0]:
        cell.border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))
        
    ws['A6'] = "En SEQUITECSA S.R.L. nos interesa saber la satisfacción de nuestros clientes, es parte de nuestros compromisos con nuestro Sistema Integrado de Gestión. Por eso nos interesa su opinión, con respecto a los servicios y/o productos brindados por nuestra empresa."
    ws.merge_cells('A6:Q6')
    ws['A6'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 40

    ws['A7'] = "DATOS DE CLIENTE"
    ws.merge_cells('A7:Q7')
    ws['A7'].alignment = Alignment(wrap_text=True)
    
    ws['A8'] = "CLIENTE"
    ws.merge_cells('A8:D8')
    ws['E8'] = cliente  # Agregar el nombre del cliente a la celda A8

    ws['A9'] = "PROYECTO"
    ws.merge_cells('A9:D9')
    ws['E9'] = proyecto

    ws['A10'] = "CONTACTO"
    ws.merge_cells('A10:D10')
    ws['E10'] = contacto

    ws['A11'] = "CARGO"
    ws.merge_cells('A11:D11')
    ws['E11'] =cargo


    ws['A12'] = "FECHA"
    ws.merge_cells('A12:D12')
    ws['E12'] = datetime.now().strftime('%Y-%m-%d')




    # Guardar el libro de trabajo en un búfer de memoria
    buffer = BytesIO()
    wb.save(buffer)

    # Crear una respuesta HTTP para el archivo Excel
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={cliente}_excel_personalizado.xlsx'

    return response

if __name__ == '__main__':
    app.run(debug=True)
